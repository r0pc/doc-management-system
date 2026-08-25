"""Training-data exporter with a hard two-key gate on real document text.

Default (synthetic-safe) mode reads manifest.csv emitted by generate_synthetic_corpus.py,
extracts each record's text via the same rendering libraries (docx/xlsx read back; PDF rows
use the manifest.json sidecar source_text, avoiding a PDF-reader dependency), and writes
dataset.csv rows: text_excerpt (first 4000 chars), label_doc_type, label_level, source.

Real-text mode (--allow-real-text <REAL_DOCUMENT_STORE_PATH>) additionally includes real
documents from the deployment store. It requires BOTH the flag AND the environment variable
DMS_EXPORT_REAL_TEXT_CONFIRM=yes. Either one alone refuses with exit code 2 and a loud
warning citing the self-hosting invariant: document text must not leave the deployment,
and Kaggle is a third-party service.
"""

from __future__ import annotations

import argparse
import csv
import json
import os
import sys
from pathlib import Path

from docx import Document
from openpyxl import load_workbook

CONFIRM_ENV = "DMS_EXPORT_REAL_TEXT_CONFIRM"
EXCERPT_CHARS = 4000
DATASET_COLUMNS = ("text_excerpt", "label_doc_type", "label_level", "source")
SUPPORTED_REAL_EXTS = frozenset({".docx", ".xlsx"})


def _refuse(order: str) -> None:
    print("=" * 72)
    print("REFUSING TO EXPORT REAL DOCUMENT TEXT")
    print()
    print("The self-hosting invariant forbids document text leaving the deployment.")
    print("Kaggle is a third-party service; uploaded text cannot be recalled.")
    print("Real-text mode needs BOTH of:")
    print("  --allow-real-text <REAL_DOCUMENT_STORE_PATH>")
    print(f"  environment variable {CONFIRM_ENV}=yes")
    print(f"Gate failure (observed order): {order}")
    print("=" * 72)
    raise SystemExit(2)


def _extract_docx(path: Path) -> str:
    return "\n".join(p.text for p in Document(path).paragraphs)


def _extract_xlsx(path: Path) -> str:
    sheet = load_workbook(path).active
    return "\n".join(
        str(cell) for row in sheet.iter_rows(values_only=True) for cell in row if cell is not None
    )


def _load_sidecar(manifest_path: Path) -> dict:
    sidecar_path = manifest_path.parent / "manifest.json"
    if not sidecar_path.exists():
        print(f"missing sidecar {sidecar_path}; regenerate the corpus first", file=sys.stderr)
        raise SystemExit(1)
    return json.loads(sidecar_path.read_text(encoding="utf-8"))


def _synthetic_rows(manifest_path: Path) -> list[dict[str, str]]:
    sidecar = _load_sidecar(manifest_path)
    text_by_version = {r["version_id"]: r["source_text"] for r in sidecar["records"]}
    rows: list[dict[str, str]] = []
    with manifest_path.open(newline="", encoding="utf-8") as fh:
        for record in csv.DictReader(fh):
            file_path = manifest_path.parent / record["file_path"]
            suffix = file_path.suffix.lower()
            if suffix == ".docx":
                text = _extract_docx(file_path)
            elif suffix == ".xlsx":
                text = _extract_xlsx(file_path)
            elif suffix == ".pdf":
                # No PDF reader dependency: use the generation-time source text.
                text = text_by_version[record["version_id"]]
            else:
                print(f"unsupported rendering {file_path.name}; skipped", file=sys.stderr)
                continue
            rows.append(
                {
                    "text_excerpt": text[:EXCERPT_CHARS],
                    "label_doc_type": record["label_doc_type"],
                    "label_level": record["label_level"],
                    "source": "synthetic",
                }
            )
    return rows


def _real_rows(real_dir: Path) -> list[dict[str, str]]:
    labels_path = real_dir / "labels.csv"
    labels: dict[str, tuple[str, str]] = {}
    if labels_path.exists():
        with labels_path.open(newline="", encoding="utf-8") as fh:
            for row in csv.DictReader(fh):
                labels[row["file_name"]] = (row["label_doc_type"], row["label_level"])

    rows: list[dict[str, str]] = []
    skipped = 0
    for path in sorted(p for p in real_dir.rglob("*") if p.suffix.lower() in SUPPORTED_REAL_EXTS):
        label = labels.get(path.name)
        if label is None:
            skipped += 1
            continue
        extract = _extract_docx if path.suffix.lower() == ".docx" else _extract_xlsx
        rows.append(
            {
                "text_excerpt": extract(path)[:EXCERPT_CHARS],
                "label_doc_type": label[0],
                "label_level": label[1],
                "source": "real",
            }
        )
    if skipped:
        print(f"{skipped} real file(s) without a labels.csv entry were skipped", file=sys.stderr)
    return rows


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="Export training rows from the synthetic corpus.")
    parser.add_argument("manifest", type=Path, help="path to manifest.csv")
    parser.add_argument("--out", type=Path, required=True, help="output dataset.csv path")
    parser.add_argument(
        "--allow-real-text",
        type=Path,
        default=None,
        metavar="REAL_DOCUMENT_STORE_PATH",
        help="include real documents from this deployment-store path (double-gated)",
    )
    args = parser.parse_args(argv)

    flag_given = args.allow_real_text is not None
    env_value = os.environ.get(CONFIRM_ENV)
    env_confirmed = env_value == "yes"

    if flag_given and not env_confirmed:
        _refuse("--allow-real-text given but confirmation env var missing or not exactly 'yes'")
    if not flag_given and env_value is not None:
        _refuse(f"{CONFIRM_ENV} set but --allow-real-text not given")

    rows = _synthetic_rows(args.manifest)
    if flag_given:
        rows.extend(_real_rows(args.allow_real_text))

    with args.out.open("w", newline="", encoding="utf-8") as fh:
        writer = csv.DictWriter(fh, fieldnames=list(DATASET_COLUMNS))
        writer.writeheader()
        writer.writerows(rows)
    sources = {row["source"] for row in rows}
    print(f"Wrote {len(rows)} row(s) to {args.out} (sources: {', '.join(sorted(sources))})")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
